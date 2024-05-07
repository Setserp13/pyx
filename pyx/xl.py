from openpyxl import Workbook, load_workbook
#from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet import worksheet

def Worksheet(): return Workbook().worksheets[0]
#def Worksheet(parent=None, title=None): return worksheet.Worksheet(parent, title) #you can't style it that way

def EmptyWorkbook():
	wb = Workbook()
	wb.remove(wb.worksheets[0])
	return wb

import math

from pyx.array_utility import *
from pyx.mat.rect import contains2
from pyx.mat.mat import to_number
import pyx.array_utility as au

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Border, Color, colors, Font, PatternFill



thick = Side(border_style='thick', color="FF000000")
thick_border = Border(left=thick, top=thick, right=thick, bottom=thick)
thin = Side(border_style='thin', color="FF000000")
thin_border = Border(left=thin, top=thin, right=thin, bottom=thin)


"""def letter_to_number(str):
	result = 0
	for i in range(len(str)):
		idx = index_of("ABCDEFGHIJKLMNOPQRSTUVWXYZ", str[i])
		result = result + idx * math.pow(26, i) if idx > -1 else print('Invalid str')#None
	return int(result)"""

"""def letter_to_number(str):
	result = 0
	for i in range(len(str)):
		try:
			result += index_of("ABCDEFGHIJKLMNOPQRSTUVWXYZ", str[i]) * math.pow(26, i)
		except:
			print('Invalid str')
	return int(result)"""


def cell_coord(address):
	for i in range(len(address)):
		if address[i].isnumeric():
			return [int(address[i:]), to_number(address[0:i]) + 1]#letter_to_number(address[0:i]) + 1]

def range_coord(address):
	min_cell, max_cell = address.split(":")
	return cell_coord(min_cell) + cell_coord(max_cell)

#from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

def cell_address(*coord): #[row, col]
	return get_column_letter(coord[1]) + str(coord[0])

def range_address(*coord): #[min_row, min_col, max_row, max_col]
	#print(coord)
	return cell_address(coord[0], coord[1]) + ":" + cell_address(coord[2], coord[3])



def add(a, b):
	c = []
	for i in range(len(a)):
		c.append(a[i] + b[i])
	return c

def translate_cell_address(address, *translation): #[row_translation, col_translation]
	return cell_address(*add(cell_coord(address), translation))

def translate_range_address(address, *translation): #[row_translation, col_translation]
	return range_address(*add(range_coord(address), translation + translation))

def translate_address(address, *translation): #[row_translation, col_translation]
	if ':' in address: return translate_range_address(address, *translation)
	return translate_cell_address(address, *translation)


#CELL ADDRESS, RANGE ADDRESS, MULTI ADDRESS




def set_row(ws, row, arr, min_col=1, ignore_merge=True, **style): #it considers merged cells
	for i in range(len(arr)):
		try:
			ws.cell(row, min_col+i).value = arr[i]
			set_cell_style(ws.cell(row, min_col+i), **style)
		except:
			print('AttributeError: \'MergedCell\' object attribute \'value\' is read-only')
		if not ignore_merge:
			min_col += merge_area(ws, row, min_col)[3] - merge_area(ws, row, min_col)[1]
			
	"""for i in range(len(arr)):
		try:
			ws.cell(row, i + min_col).value = arr[i]
		except:
			print('AttributeError: \'MergedCell\' object attribute \'value\' is read-only')
		if not ignore_merge:
			min_col += merge_area(ws, row, min_col)[3] - merge_area(ws, row, min_col)[1]"""

def set_col(ws, col, arr, min_row=1):
	for i in range(len(arr)): ws.cell(i + min_row, col).value = arr[i]

def set_rng(ws, mx, min_row=1, min_col=1, **style):
	for i in range(len(mx)):
		for j in range(len(mx[i])):
			ws.cell(min_row+i, min_col+j).value = mx[i][j]
			set_cell_style(ws.cell(min_row+i, min_col+j), **style)

def rng_to_list(ws, min_row, min_col, max_row, max_col):
	result = []
	for i in range(min_row, max_row + 1):
		for j in range(min_col, max_col + 1):
			result.append(ws.cell(i, j).value)
	return result

	
###def horizontal_cell_stretch(ws, row, col, scale):
	

from copy import copy

def stylecpy(dst, src):
	if src.has_style: dst._style = copy(src._style)


from openpyxl.formula.translate import Translator

def cellcpy(dst, src):
	if src.data_type == 'f':
		dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
	else:
		dst.value = src.value
	#style
	for k in ['alignment', 'border', 'fill', 'font', 'number_format', 'protection']:
		setattr(dst, k, copy(getattr(src, k)))


from copy import deepcopy
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange



def concatws(objs):
	result = Worksheet()
	for x in objs:
		wscpy(result, x, min_row=result.max_row)
	return result


"""def foreach(ws, action, min_row=1, min_col=1, max_row=None, max_col=None):
	if max_row == None:
		max_row = ws.max_row
	elif max_col == None:
		max_col = ws.max_col
	for i in range(1, max_row + 1):
		for j in range(1, max_column + 1):
			action(ws.cell(i, j), i, j)"""

def wscpy(dst, src, min_row=1, min_col=1):
	for i in range(0, src.max_row):
		for j in range(0, src.max_column):
			cellcpy(dst.cell(i + min_row, j + min_col), src.cell(i + 1, j + 1))
			#dst.cell(i + min_row, j + min_col).value = src.cell(i + 1, j + 1).value
			#stylecpy(dst.cell(i + min_row, j + min_col), src.cell(i + 1, j + 1))
	for merged_cell in src.merged_cells.ranges:
		#print(merged_cell)
		if ":" not in str(merged_cell): continue
		dst.merge_cells(translate_range_address(str(merged_cell), min_row - 1, min_col - 1))
	for dv in src.data_validations.dataValidation:
		dvcopy = deepcopy(dv)
		dvcopy.sqref = MultiCellRange()
		dst.add_data_validation(dvcopy)
		for address in str(dv.sqref).split(' '):
			dvcopy.add(translate_address(address, min_row - 1, min_col - 1))
		#print(dvcopy)


def dfcpy(ws, df, min_row=1, min_col=1, header_outline=thick, header_inline=thin, outline=thick, inline=thin, **style):
	set_row(ws, min_row, df.columns, min_col, **{ k.replace('header_', ''):style[k] for k in style if k.startswith('header_') })
	set_rng(ws, df.values, min_row + 1, min_col, **{ k:style[k] for k in style if not k.startswith('header_') })
	set_range_border(ws, min_row, min_col, min_row, min_col + len(df.columns) - 1, header_outline, header_inline)
	set_range_border(ws, min_row + 1, min_col, min_row + df.shape[0], min_col + len(df.columns) - 1, outline, inline)
	#dfvalscpy(ws, df, min_row + 1, min_col)


from pyx.array_utility import call

def map(df, ws, func, min_row=1, min_col=1):
	for i in range(df.shape[0]):
		for j in range(df.shape[1]):
			ws.cell(i + min_row, j + min_col).value = call(func, df.iat[i, j], i, j)

from pandas import DataFrame

DataFrame.map = map



def dfvalscpy(ws, df, min_row=1, min_col=1):
	df.map(ws, lambda x, i, j: x, min_row, min_col)

def dfvalscpytransp(ws, df, min_row=1, min_col=1):
	df.map(ws, lambda x, i, j: df.iat[i, j], min_row, min_col)




from openpyxl.utils.cell import range_boundaries


def merge_areas(ws, min_row, min_col, max_row, max_col):
	result = []
	for range in ws.merged_cells.ranges:
		boundaries = range_boundaries(str(range))
		if contains2([min_row, min_col], [max_row, max_col], [boundaries[1], boundaries[0]]) or contains2([min_row, min_col], [max_row, max_col], [boundaries[3], boundaries[2]]):
			result.append([boundaries[1], boundaries[0], boundaries[3], boundaries[2]])
	return result

def merge_area(ws, row, col):
	for range in ws.merged_cells.ranges:
		boundaries = range_boundaries(str(range))
		if contains2([boundaries[1], boundaries[0]], [boundaries[3], boundaries[2]], [row, col]):
			return [boundaries[1], boundaries[0], boundaries[3], boundaries[2]]
	return [row, col, row, col]

def is_merged(ws, row, col):
	ma = merge_area(ws, row, col)
	return ma[0] != ma[2] or ma[1] != ma[3]

def copy_merge(ws, range_coord, min_row, min_col):
	for i, x in enumerate(range(range_coord[0], range_coord[2] + 1)):
		for j, y in enumerate(range(range_coord[1], range_coord[3] + 1)):
			if is_merged(ws, min_row + i, min_col + j):
				continue
			ma = merge_area(ws, x, y)
			ws.merge_cells(range_address(min_row + i, min_col + j, min_row + i + (ma[2] - ma[0]), min_col + j + (ma[3] - ma[1])))
			



def get_cell(ws, i, j):
	area = merge_area(ws, i, j)
	return ws.cell(area[0], area[1])


def get_value(ws, i, j, dflt=None):
	return get_cell(ws, i, j).value if i > 0 and j > 0 and i <= ws.max_row and j <= ws.max_column else dflt



def groups_in_column(ws, col, min_row, max_row):
	result = [] #contains items like [group_min_row, group_min_row]
	index = -1
	for i in range(min_row, max_row + 1):
		if index == -1:
			result.append([i, i])
			index = index + 1
		else:
			#if ws.cell(i, col).value == ws.cell(result[index][0], col).value
			if get_value(ws, i, col) == get_value(ws, result[index][0], col): #ws.cell(result[index][0], col).value
				result[index][1] = i
			else:
				result.append([i, i])
				index = index + 1
	return result



def merge_vertically(ws, min_row, min_col, max_row, max_col):
	for j in range(min_col, max_col + 1):
		groups = groups_in_column(ws, j, min_row, max_row)
		for group in groups:
			ws.merge_cells(range_address(group[0], j, group[1], j))







def for_each_cell(ws, start_row, start_column, end_row, end_column, action):
	for i in range(start_row, end_row + 1):
		for j in range(start_column, end_column):
			action(ws.cell(i, j))

from openpyxl.styles import Alignment

def set_content_alignment(ws, h = 'center', v = 'center'):
	for i in range(1, ws.max_row + 1):
		for j in range(1, ws.max_column + 1):
			if ws.cell(i, j).value != None:
				ws.cell(i, j).alignment = Alignment(horizontal=h, vertical=v)



def merge_workbooks(arr):
	wb = Workbook()
	for i in range(len(arr)):
		add_worksheets(wb, arr[i].worksheets)        
	return wb

def add_worksheet(wb, ws):
	return wscpy(wb.create_sheet(ws.title), ws)



def add_worksheets(wb, arr):
	foreach(arr, lambda x: add_worksheet(wb, x))
	"""result = []
	for i in range(len(arr)):
		result.append(add_worksheet(wb, arr[i]))    
	return result"""


def autofit_column(ws, index):
	ws.column_dimensions[get_column_letter(index)].bestFit = True
	ws.column_dimensions[get_column_letter(index)].auto_size = True
	#ws.column_dimensions[get_column_letter(index)].width = reduce(list(map(ws[get_column_letter(index)], lambda x, i: len(str(x.value)))), lambda total, currentValue: max(total, currentValue))


def autofit_all_columns(ws):
	for i in range(1, ws.max_column): autofit_column(ws, i)


def set_column_width(ws, index, value = 10):
	ws.column_dimensions[get_column_letter(index)].width = value

def set_columns_width(ws, value = 10):
	if isinstance(value, list):
		for i, x in enumerate(value):
			set_column_width(ws, i + 1, x)
	else:
		for i in range(1, ws.max_column + 1):
			set_column_width(ws, i, value)

















def total_row(min_row, min_col, max_row, max_col, formula="SUM"):
	result = []
	for j in range(min_col, max_col + 1):
		result.append("=" + formula + "(" + range_address(min_row, j, max_row, j) + ")")
	return result











"""import copy
copy.copy(x) shallow copy
copy.deepcopy(x) deep copy"""










#def size(ws): return [0, 0] if ws == None else [ws.max_row, ws.max_column]






import pyx.matrix_utility as mu

def max_w(matrices):
	return mu.map_columns(matrices, lambda x, i: reduce([y.max_column for y in x], lambda total, currentValue: max(total, currentValue)))

def max_h(matrices):
	return [reduce([y.max_row for y in x], lambda total, currentValue: max(total, currentValue)) for x in matrices]



def min_cells(matrices):
	mw = max_w(matrices)
	mh = max_h(matrices)
	result = mu.full(len(mh), len(mw), None)#, fill_value = [0, 0])
	current_start_row = 0     
	for i in range(0, len(result)):
		current_start_column = 0
		for j in range(0, len(result[i])):
			result[i][j] = [current_start_row + 1, current_start_column + 1]
			current_start_column += mw[j]
		current_start_row += mh[i]
	return result

def paste_worksheet_matrix(ws, worksheet_matrix, spacing = 1):
	sc = min_cells(worksheet_matrix)
	for i in range(len(worksheet_matrix)):
		for j in range(len(worksheet_matrix[i])):
			sc[i][j][0] += i * spacing
			sc[i][j][1] += j * spacing
			wscpy(ws, worksheet_matrix[i][j], sc[i][j][0], sc[i][j][1])









def set_cell_style(cell, **style):
	for k in style:
		setattr(cell, k, style[k])

def set_range_style(ws, min_row, min_col, max_row, max_col, **style):
	for i in range(min_row, max_row + 1):
		for j in range(min_col, max_col + 1):
			set_cell_style(ws.cell(i, j), **style)

def banded_rows(ws, min_row, min_col, max_row, max_col, colors=['FFFFFFFF', 'FFBFBFBF']):
	for i in range(min_row, max_row + 1):
		set_range_style(ws, i, min_col, i, max_col, fill=color_fill(colors[i % len(colors)]))


def merge_range(ws, min_row, min_col, max_row, max_col, value, **style):
	ws.merge_cells(range_address(min_row, min_col, max_row, max_col))
	ws.cell(min_row, min_col).value = value #set upper left cell value
	#set_cell_style(ws.cell(min_row, min_col), **style)
	set_range_style(ws, min_row, min_col, max_row, max_col, **style)

def color_fill(color): return PatternFill(start_color=color, end_color=color, fill_type='solid')
	



def set_range_border(ws, min_row, min_col, max_row, max_col, outline=thick, inline=thin):
	for i in range(min_row, max_row + 1):
		for j in range(min_col, max_col + 1):
			border = Border(left=inline, top=inline, right=inline, bottom=inline)
			area = merge_area(ws, i, j)
			"""if i == min_row: border.top = outline
			if j == min_col: border.left = outline
			if i == max_row: border.bottom = outline
			if j == max_col: border.right = outline"""
			if area[0] == min_row: border.top = outline
			if area[1] == min_col: border.left = outline
			if area[2] == max_row: border.bottom = outline
			if area[3] == max_col: border.right = outline
			ws.cell(i, j).border = border



def set_content_border(ws, outline=thick, inline=thin):
	for i in range(1, ws.max_row + 1):
		for j in range(1, ws.max_column + 1):

			area = merge_area(ws, i, j)

			if ws.cell(area[0], area[1]).value == None: continue


			#if ws.cell(i, j).value == None: continue
			border = Border(left=inline, top=inline, right=inline, bottom=inline)
			if get_value(ws, i - 1, j) == None: border.top = outline
			if get_value(ws, i, j - 1) == None: border.left = outline
			if get_value(ws, i + 1, j) == None: border.bottom = outline
			if get_value(ws, i, j + 1) == None: border.right = outline
			ws.cell(i, j).border = border

def set_range_border_by_group(ws, min_row, min_col, max_row, max_col, outline=thick, inline=thin):
	i = min_row
	while i < max_row + 1:
		area = merge_area(ws, i, min_col)
		#print(area)
		set_range_border(ws, i, min_col, area[2], max_col, outline, inline)
		i = area[2] + 1

def replace(ws, old_value, new_value):
	for i in range(1, ws.max_row + 1):
		for j in range(1, ws.max_column + 1):
			if isinstance(ws.cell(i, j).value, str):
				ws.cell(i, j).value = ws.cell(i, j).value.replace(old_value, new_value)

